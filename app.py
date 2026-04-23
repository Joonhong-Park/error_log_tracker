from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from datetime import datetime
import io
import os
import sys
import webbrowser
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import db
import pg_sync

# PyInstaller 번들 실행 시 templates 경로 보정
if getattr(sys, "frozen", False):
    _base = sys._MEIPASS
else:
    _base = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__, template_folder=os.path.join(_base, "templates"))
app.secret_key = "error_logging_secret"

db.init_db()


@app.route("/")
def index():
    show_resolved = request.args.get("show_resolved", "0") == "1"
    date_from     = request.args.get("date_from", "")
    date_to       = request.args.get("date_to", "")
    table_name    = request.args.get("table_name", "")
    load_type     = request.args.get("load_type", "")
    error_search  = request.args.get("error_search", "")
    cause_search  = request.args.get("cause_search", "")
    sort_by       = request.args.get("sort_by", "create_date_ts")
    sort_dir      = request.args.get("sort_dir", "desc")
    try:
        page = max(1, int(request.args.get("page", 1)))
    except (ValueError, TypeError):
        page = 1
    per_page      = 50

    rows, total = db.get_list(
        show_resolved=show_resolved,
        date_from=date_from or None,
        date_to=date_to or None,
        table_name=table_name or None,
        load_type=load_type or None,
        error_search=error_search or None,
        cause_search=cause_search or None,
        sort_by=sort_by,
        sort_dir=sort_dir,
        page=page,
        per_page=per_page,
    )
    total_pages = max(1, (total + per_page - 1) // per_page)
    load_types = db.get_filter_options()

    return render_template(
        "index.html",
        rows=rows,
        show_resolved=show_resolved,
        date_from=date_from,
        date_to=date_to,
        table_name=table_name,
        load_type=load_type,
        error_search=error_search,
        cause_search=cause_search,
        sort_by=sort_by,
        sort_dir=sort_dir,
        page=page,
        total=total,
        total_pages=total_pages,
        per_page=per_page,
        load_types=load_types,
    )


@app.route("/sync")
def sync():
    try:
        total, added = pg_sync.sync()
        flash(f"동기화 완료 — 전체 {total}건 중 신규 {added}건 추가", "success")
    except Exception as e:
        flash(f"동기화 실패: {e}", "error")
    return redirect(url_for("index"))


@app.route("/error/<message_id>", methods=["GET", "POST"])
def detail(message_id):
    row = db.get_one(message_id)
    if row is None:
        flash("해당 오류 건을 찾을 수 없습니다.", "error")
        return redirect(url_for("index"))

    if request.method == "POST":
        root_cause = request.form.get("root_cause", "").strip()
        action_required = request.form.get("action_required", "").strip()
        action_taken = request.form.get("action_taken", "").strip()
        resolved = request.form.get("resolved") == "1"
        db.update_memo(message_id, root_cause, action_required, action_taken, resolved)
        flash("저장되었습니다.", "success")
        return redirect(url_for("index"))

    return render_template("detail.html", row=row)


@app.route("/export")
def export():
    show_resolved = request.args.get("show_resolved", "0") == "1"
    date_from     = request.args.get("date_from", "") or None
    date_to       = request.args.get("date_to", "") or None
    table_name    = request.args.get("table_name", "") or None
    load_type     = request.args.get("load_type", "") or None
    error_search  = request.args.get("error_search", "") or None
    cause_search  = request.args.get("cause_search", "") or None
    sort_by       = request.args.get("sort_by", "create_date_ts")
    sort_dir      = request.args.get("sort_dir", "desc")
    rows, _       = db.get_list(
        show_resolved=show_resolved, date_from=date_from, date_to=date_to,
        table_name=table_name, load_type=load_type, error_search=error_search,
        cause_search=cause_search, sort_by=sort_by, sort_dir=sort_dir, per_page=None,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "오류 목록"

    # 헤더 스타일
    header_font    = Font(bold=True, color="FFFFFF", size=10)
    header_fill    = PatternFill("solid", fgColor="1E3A5F")
    header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    cell_align     = Alignment(vertical="top", wrap_text=True)

    headers = ["발생 시각", "파일명", "테이블명", "오류 내용", "원인 파악", "조치 필요 내용", "실제 조치 내용", "조치 여부", "완료 일시"]
    col_widths = [20, 30, 20, 50, 35, 35, 35, 12, 20]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = header_align
        cell.border  = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22

    # 데이터 행
    resolved_fill   = PatternFill("solid", fgColor="D4EDDA")
    unresolved_fill = PatternFill("solid", fgColor="FFF3CD")

    for row_idx, r in enumerate(rows, start=2):
        resolved_text = "완료" if r["resolved"] else "미완료"
        values = [
            r["create_date_ts"],
            r["origin_file_name"],
            r["table_name"],
            r["error"],
            r["root_cause"] or "",
            r["action_required"] or "",
            r["action_taken"] or "",
            resolved_text,
            r["resolved_date_ts"] or "",
        ]
        row_fill = resolved_fill if r["resolved"] else unresolved_fill

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border    = thin_border
            # 조치여부 컬럼만 배경색 적용
            if col_idx == 8:
                cell.fill = row_fill

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # 파일명에 날짜 포함
    filename = f"nifi_error_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    import threading
    threading.Timer(1.0, lambda: webbrowser.open("http://localhost:5000")).start()
    app.run(debug=False, host="127.0.0.1", port=5000)
